from copy import copy
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .cheat_sheet import CheatSheet, SheetLine
from .session import Session
from .utilities import get_delta_time_HM, short_event, short_gender

# excel worksheet board style
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# excel worksheet font style
_FONT_STYLE = Font(size="14")
_FONT_STYLE_BOLD = Font(size="14", bold=True)

# excel worksheet column widths
_BUFFER_COLUMN_WIDTH: int = 6
_EVENT_COLUMN_WIDTH: int = 6
_NAME_COLUMN_WIDTH: int = 16


class ExcelWorkbook:
    def __init__(self) -> None:
        self.wb: Workbook = Workbook()

    def write(self, filename: str) -> None:
        # We don't need or want the default worksheet.  It's easier to just
        # delete it than renaming it.  Based off of user input, there may be
        # a condition where we don't create any worksheets and that causes
        # a exception when saving.  To prevent this we're only deleting the
        # default sheet if we have created others.
        if len(self.wb.worksheets) > 1:
            self.wb.remove(self.wb["Sheet"])

        # Save the file
        self.wb.save(filename)

        # TODO: Needs try catch

    def _get_line(self, womens: str, mens: str) -> str:
        """Combines women/mens events and heats in three column mode.
        If there is a single gender, a seperator is not needed"""

        seperator: str = "/"

        if womens == "" or mens == "":
            seperator = ""

        return womens + seperator + mens

    def _create_unique_sheet(self, name: str) -> Worksheet:
        """Sheets are typicalled named based off of the session number.  To prevent
        duplicating names during testing, this function creates a unique tab name based
        on the initial 'name' passed in

            Args:
                name: Initial tab name

            Return:
                Correctly named worksheet in the workbook
        """
        sheet_name: str = name
        counter: int = 1

        while sheet_name in self.wb.sheetnames:
            sheet_name = f"{name}.{counter}"
            counter += 1

        return self.wb.create_sheet(sheet_name)

    def _cheat_sheet_setup(self, s: Session, three_column: bool = False) -> Worksheet:
        """Creates worksheet, sets up the column widths and adds headers"""
        ws: Worksheet = self._create_unique_sheet(f"S{s.number}")

        if three_column:
            ws.column_dimensions[get_column_letter(1)].width = _EVENT_COLUMN_WIDTH * 2
            ws.column_dimensions[get_column_letter(2)].width = _NAME_COLUMN_WIDTH
            ws.column_dimensions[get_column_letter(3)].width = _EVENT_COLUMN_WIDTH * 2
            ws.column_dimensions[get_column_letter(4)].width = _BUFFER_COLUMN_WIDTH
        else:
            ws.column_dimensions[get_column_letter(1)].width = _EVENT_COLUMN_WIDTH
            ws.column_dimensions[get_column_letter(2)].width = _EVENT_COLUMN_WIDTH
            ws.column_dimensions[get_column_letter(3)].width = _NAME_COLUMN_WIDTH
            ws.column_dimensions[get_column_letter(4)].width = _EVENT_COLUMN_WIDTH
            ws.column_dimensions[get_column_letter(5)].width = _EVENT_COLUMN_WIDTH
            ws.column_dimensions[get_column_letter(6)].width = _BUFFER_COLUMN_WIDTH

        # header
        if three_column:
            ws.cell(row=1, column=1, value="W/M")
            ws.cell(row=1, column=2, value=f"Session {s.number}")
            ws.cell(row=1, column=3, value="W/M")
        else:
            ws.cell(row=1, column=1, value="W")
            ws.cell(row=1, column=2, value="H")
            ws.cell(row=1, column=3, value=f"Session {s.number}")
            ws.cell(row=1, column=4, value="M")
            ws.cell(row=1, column=5, value="H")

        return ws

    def _write_cheat_sheet_line(
        self, ws: Worksheet, three_column: int, index: int, line: SheetLine
    ) -> None:
        """Writes a single event (women and men) to a cheat sheet

        Args:
            ws: Worksheet reference
            three_column: Boolean flag whether this is a three or five
                          column cheat sheet
            index: Row intex to write to
            line: line containing event/heat/strok/distance
        """

        if three_column:
            ws.cell(
                row=index,
                column=1,
                value=self._get_line(str(line.womems_ev_num), str(line.mens_event_num)),
            )
            ws.cell(row=index, column=2, value=line.event_name)
            ws.cell(
                row=index,
                column=3,
                value=self._get_line(
                    str(line.womens_heat_count), str(line.mens_heat_count)
                ),
            )

        else:
            ws.cell(row=index, column=1, value=line.womems_ev_num)
            ws.cell(row=index, column=2, value=line.womens_heat_count)
            ws.cell(row=index, column=3, value=line.event_name)
            ws.cell(row=index, column=4, value=line.mens_event_num)
            ws.cell(row=index, column=5, value=line.mens_heat_count)

    def make_cheat_sheets(
        self, sessions: list[Session], three_column: bool = False
    ) -> None:
        """Takes a list of sessions and returns an MS Excel worksboot with a single tab per session.

        Args:
            sessions: List of sessions to be added to the XLS.  If tagged "merged," the sessions will
                      not be added to the XLS.
            three_column: Boolean flag whether a three (True) or Five (False) column sheet should be built
        """

        max_column: int = 3 if three_column else 5

        # do not process the original sessions that have been merged
        filtered_sessions = (s for s in sessions if not s.has_been_merged)

        for s in filtered_sessions:
            row_count = 1

            ws: Worksheet = self._cheat_sheet_setup(s, three_column)

            cs: CheatSheet = CheatSheet(s)

            # all sessions
            for index, line in enumerate(cs.lines, start=2):
                row_count += 1

                # notes are on merged lines
                if line.is_note:
                    ws.merge_cells(
                        start_row=index,
                        start_column=1,
                        end_row=index,
                        end_column=max_column,
                    )
                    ws.cell(row=index, column=1, value=line.event_name)

                else:
                    self._write_cheat_sheet_line(ws, three_column, index, line)

                # add border
                for c in range(1, (max_column + 1)):
                    # header row
                    ws.cell(row=1, column=c).font = _FONT_STYLE_BOLD
                    ws.cell(row=1, column=c).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

                    # event and note rows
                    for r in range(2, row_count + 1):
                        ws.cell(row=r, column=c).border = _THIN_BORDER
                        ws.cell(row=r, column=c).alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

                        ws.cell(row=r, column=c).font = _FONT_STYLE

                # Make copies of the cheat sheet assembled above two more times.
                start: int = 1
                end: int = max_column + 1
                for i in range(2):
                    ExcelWorkbook.copy_columns(ws, start, end, end + 1)
                    start = end + 1
                    end += max_column + 1

    def make_event_durations(self, sessions: list[Session]):
        """Prints seperate report tab that gives timing
        information for every event in each session"""

        unmerged_sessions = (s for s in sessions if not s.merged_session)

        fill_yellow = PatternFill(
            start_color="FFFFCC", end_color="FFFFCC", fill_type="lightDown"
        )

        fill_blue = PatternFill(
            start_color="DDEBF7", end_color="DDEBF7", fill_type="lightDown"
        )

        ws = self._create_unique_sheet("Duration")

        # set column widths
        header_widths: list[int] = [10, 10, 50]
        for j in range(len(header_widths)):
            ws.column_dimensions[get_column_letter(j + 1)].width = header_widths[j]

        # 'row' always points to the first row to be written to
        row: int = 1
        for s in unmerged_sessions:
            ws.cell(row=row, column=1, value=f"\nSession {s.number}")
            ws.cell(row=row, column=1).font = Font(size="14", bold=True)
            row += 1

            row_offset: int = row

            for i in range(len(s.events)):
                start_time = s.events[i].datetime_start

                if i > 0 and s.events[i - 1].break_follows:
                    ws.cell(
                        row=row_offset + i,
                        column=1,
                        value=f"{s.events[i - 1].break_time}-min",
                    )
                    ws.cell(row=row_offset + i, column=2, value="")
                    ws.cell(row=row_offset + i, column=3, value="BREAK")
                    ws.cell(row=row_offset + i, column=3).fill = fill_blue

                else:
                    # use session end for the final event stop time
                    if i == len(s.events) - 1:
                        finish_time: datetime = s.datetime_finish
                    else:
                        finish_time = s.events[i + 1].datetime_start

                    event_duration: timedelta = finish_time - start_time
                    delta_in_minutes: int = round(event_duration.total_seconds() / 60)

                    line1: str = f"{delta_in_minutes}-min "
                    line2: str = f"h={s.events[i].heat_count}"
                    line3: str = (
                        f"E{s.events[i].number} "
                        + f"{short_gender(s.events[i].gender)} "
                        + f"{s.events[i].distance} "
                        + f"{short_event(s.events[i].stroke)}"
                        + f"{' Relay' if s.events[i].is_relay else ''}"
                    )

                    ws.cell(row=row_offset + i, column=1, value=line1)
                    ws.cell(row=row_offset + i, column=2, value=line2)
                    ws.cell(row=row_offset + i, column=3, value=line3)

                    if s.events[i].is_relay:
                        ws.cell(row=row_offset + i, column=3).fill = fill_yellow

                row += 1

            # add a single row space before each column
            row += 1

    def make_session_diagnostics(self, sessions: list[Session]):
        """Prints seperate report tab that gives timing and pool information
        per sessoion"""

        unmerged_sessions = (s for s in sessions if not s.merged_session)

        ws = self._create_unique_sheet("Diagnostics")

        font_style_header = Font(bold=True)

        red_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )

        header_names: list[str] = [
            "Session",
            "Lanes",
            "Heats",
            "Splashes",
            "Interval",
            "+ Back",
            "Relays",
            "Breaks",
            "Time",
            "12U Finish",
            "12U End Event",
        ]
        header_widths: list[int] = [10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 20]

        # header
        for j in range(len(header_names)):
            ws.cell(1, j + 1, value=header_names[j])
            ws.column_dimensions[get_column_letter(j + 1)].width = header_widths[j]
            ws.cell(1, j + 1).font = font_style_header
            ws.cell(1, j + 1).alignment = Alignment(vertical="center")

        # write all session data
        for row, s in enumerate(unmerged_sessions, start=2):
            heats = 0
            splashes = 0
            relays = 0
            breaks = 0

            # count all session heats and entries (aka 'splashes')
            for e in s.events:
                heats += e.heat_count
                splashes += e.total_entries
                relays += e.is_relay
                breaks += e.break_follows

            (hours, minutes) = get_delta_time_HM(s.datetime_start, s.datetime_finish)

            ws.cell(row, 1, value=f"{s.number}")
            ws.cell(row, 2, value=f"{s.pool_size}")
            ws.cell(row, 3, value=f"{heats}")
            ws.cell(row, 4, value=f"{splashes}")
            ws.cell(row, 5, value=f"{s.interval}")
            ws.cell(row, 6, value=f"{s.plus_back_interval}")
            ws.cell(row, 7, value=f"{relays}")
            ws.cell(row, 8, value=f"{breaks}")
            ws.cell(row, 9, value=f"{hours}h {minutes}m")

            if s.last_12U_event is not None:
                final_12U_index: int = s.last_12U_event

                # is the final event in the session
                # if so, the final 12U is the session end time
                if final_12U_index == len(s.events) - 1:
                    end_time_12U = s.datetime_finish
                else:
                    # else, the final time is the start time of the
                    # event following the final 12U event
                    end_time_12U = s.events[final_12U_index + 1].datetime_start

                (hours, minutes) = get_delta_time_HM(s.datetime_start, end_time_12U)

                ws.cell(row, 10, value=f"{hours}h {minutes}m")
                ws.cell(
                    row,
                    11,
                    value=f"E{s.events[final_12U_index].number} "
                    f"{s.events[final_12U_index].distance} "
                    f"{short_event(s.events[final_12U_index].stroke)}",
                )

                # If the 12U finish time is greater than 4-hours, flag
                # the session as a possible 4-hour violation.  This tool
                # cannot detrmine whether this is a championship meet and
                # therefore does need to comply
                if hours >= 4:
                    ws.cell(row, 10).fill = red_fill

                    # count heats within 12U window
                    heats: int = 0
                    for i in range(final_12U_index + 1):
                        heats += s.events[i].heat_count

                    # how much time can we save dropping the interval by 5-seoncds?
                    saved_min: int = (heats * 5) // 60
                    line: str = f"12U Heats {heats}, 5-sec saves {saved_min}-min"
                    ws.cell(row, 12, value=line)

    @staticmethod
    def copy_columns(ws, start_col: int, end_col: int, destination_col: int) -> None:
        """Copy columns, including values, formatting, merged cells, and widths.
        Borders require a little bit more work to keep track off when cells are merged"""

        offset = destination_col - start_col

        # Copy column widths
        for source_col in range(start_col, end_col + 1):
            source_letter = ws.cell(row=1, column=source_col).column_letter
            destination_letter = ws.cell(
                row=1,
                column=source_col + offset,
            ).column_letter

            ws.column_dimensions[destination_letter].width = ws.column_dimensions[
                source_letter
            ].width

        # Copy merged ranges and their borders
        for merged_range in list(ws.merged_cells.ranges):
            if start_col <= merged_range.min_col and merged_range.max_col <= end_col:
                new_min_col = merged_range.min_col + offset
                new_max_col = merged_range.max_col + offset

                # Save the borders before merging.
                borders = {}

                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        borders[(row, col)] = copy(cell.border)

                # Create the merged range.
                ws.merge_cells(
                    start_row=merged_range.min_row,
                    end_row=merged_range.max_row,
                    start_column=new_min_col,
                    end_column=new_max_col,
                )

                # Restore the borders on the copied range.
                for (row, col), border in borders.items():
                    new_cell = ws.cell(
                        row=row,
                        column=col + offset,
                    )
                    new_cell.border = copy(border)

        # Copy cells
        for row in ws.iter_rows():
            for cell in row[start_col - 1 : end_col]:
                if isinstance(cell, MergedCell):
                    continue

                new_cell = ws.cell(
                    row=cell.row,
                    column=cell.column + offset,
                )

                new_cell.value = cell.value

                if cell.has_style:
                    new_cell._style = copy(cell._style)

                if cell.hyperlink:
                    new_cell._hyperlink = copy(cell.hyperlink)

                if cell.comment:
                    new_cell.comment = copy(cell.comment)
